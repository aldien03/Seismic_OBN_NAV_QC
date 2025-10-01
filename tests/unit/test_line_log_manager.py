"""
Unit tests for LineLogManager module.

Tests cover:
- Initialization and configuration
- File finding with regex patterns
- Workbook opening with retry logic
- Content generation and formatting
- Label mapping for QC checks
- Range detection for shot points
- Cell finding and formatting
- Update operations
- Shot point marker extraction (FGSP, LGSP, FASP, etc.)
"""

import pytest
import os
import re
import tempfile
import pandas as pd
from datetime import datetime
from configparser import ConfigParser
from openpyxl import Workbook
from openpyxl.styles import Alignment

from line_log_manager import LineLogManager


@pytest.fixture
def mock_config():
    """Create mock configuration for tests."""
    config = ConfigParser()
    config.add_section('LineLog')
    config.set('LineLog', 'max_open_attempts', '5')
    config.set('LineLog', 'acquisition_comments_label', 'Acquisition and Processing Comments')
    return config


@pytest.fixture
def line_log_manager(mock_config):
    """Create LineLogManager instance for tests."""
    return LineLogManager(mock_config)


@pytest.fixture
def sample_merged_df():
    """Create sample merged DataFrame for testing."""
    return pd.DataFrame({
        'shot_point': [1001, 1002, 1003],
        'datetime_UTC': [datetime(2025, 10, 1, 10, 0, 0),
                        datetime(2025, 10, 1, 10, 0, 6),
                        datetime(2025, 10, 1, 10, 0, 12)]
    })


@pytest.fixture
def sample_percentages():
    """Create sample percentages dictionary."""
    return {
        'percent_radial': 2.5,
        'percent_gd_errors': 1.8
    }


@pytest.fixture
def sample_log_data():
    """Create sample log data dictionary."""
    return {
        'log_sub_array_sep_flag': [1001, 1003],
        'log_gun_depth_flag': [1002],
        'log_autofires': [1005],
        'log_timing_warning': [(1001, ['G1', 'G2'])],
        'log_repeatability_flag': [1001, 1003, 1005, 1007],
        'log_consec_7_source_errors': ['1010-1024'],
    }


class TestLineLogManagerInit:
    """Test LineLogManager initialization."""

    def test_initialization(self, mock_config):
        """Test basic initialization."""
        manager = LineLogManager(mock_config)
        assert manager.config == mock_config
        assert manager.max_attempts == 5
        assert manager.comments_label == 'Acquisition and Processing Comments'

    def test_line_log_pattern_is_valid_regex(self, line_log_manager):
        """Test that line log pattern is valid regex."""
        pattern = line_log_manager.line_log_pattern
        # Should compile without error
        re.compile(pattern)

        # Test matching
        assert re.match(pattern, '0256-3184P31885_Nav_LineLog.xlsm')
        assert not re.match(pattern, '3184P31885_Nav_LineLog.xlsm')
        assert not re.match(pattern, '0256-3184P31885_LineLog.xlsx')

    def test_initialization_with_fallback_values(self):
        """Test initialization with fallback configuration values."""
        config = ConfigParser()
        config.add_section('LineLog')
        # Don't set max_open_attempts or acquisition_comments_label

        manager = LineLogManager(config)
        assert manager.max_attempts == 5  # fallback value
        assert manager.comments_label == 'Acquisition and Processing Comments'  # fallback


class TestFindLineLogFile:
    """Test find_line_log_file method."""

    def test_find_file_success(self, line_log_manager, tmp_path):
        """Test finding line log file successfully."""
        # Create a valid line log file
        line_log_file = tmp_path / "0256-3184P31885_Nav_LineLog.xlsm"
        line_log_file.touch()

        result = line_log_manager.find_line_log_file(str(tmp_path))
        assert result is not None
        assert result == str(line_log_file)

    def test_find_file_invalid_directory(self, line_log_manager):
        """Test with invalid directory."""
        result = line_log_manager.find_line_log_file('/nonexistent/directory')
        assert result is None

    def test_find_file_no_matching_file(self, line_log_manager, tmp_path):
        """Test when no matching file exists."""
        # Create non-matching files
        (tmp_path / "wrong_format.xlsm").touch()
        (tmp_path / "3184P31885_LineLog.xlsm").touch()

        result = line_log_manager.find_line_log_file(str(tmp_path))
        assert result is None

    def test_find_file_multiple_files_returns_first(self, line_log_manager, tmp_path):
        """Test that first matching file is returned when multiple exist."""
        file1 = tmp_path / "0256-3184P31885_Nav_LineLog.xlsm"
        file2 = tmp_path / "0256-3184P31886_Nav_LineLog.xlsm"
        file1.touch()
        file2.touch()

        result = line_log_manager.find_line_log_file(str(tmp_path))
        assert result is not None
        assert os.path.basename(result) in ['0256-3184P31885_Nav_LineLog.xlsm',
                                            '0256-3184P31886_Nav_LineLog.xlsm']


class TestOpenWorkbookWithRetry:
    """Test open_workbook_with_retry method."""

    def test_open_workbook_success(self, line_log_manager, tmp_path):
        """Test successfully opening a workbook."""
        # Create a valid Excel file
        wb = Workbook()
        test_file = tmp_path / "test.xlsm"
        wb.save(str(test_file))
        wb.close()

        result_wb = line_log_manager.open_workbook_with_retry(str(test_file))
        assert result_wb is not None
        result_wb.close()

    def test_open_workbook_nonexistent_file(self, line_log_manager):
        """Test opening nonexistent file."""
        result = line_log_manager.open_workbook_with_retry('/nonexistent/file.xlsm')
        assert result is None


class TestGenerateContent:
    """Test _generate_content method."""

    def test_generate_content_basic(self, line_log_manager, sample_merged_df,
                                   sample_percentages, sample_log_data):
        """Test basic content generation."""
        content = line_log_manager._generate_content(
            sample_merged_df, sample_percentages, sample_log_data, [], []
        )

        assert "Shooting Mode: 4D Source" in content
        assert "97.50%" in content  # 100 - 2.5
        assert "98.20%" in content  # 100 - 1.8

    def test_generate_content_with_missing_sp(self, line_log_manager, sample_merged_df,
                                             sample_percentages):
        """Test content generation with missing shot points."""
        content = line_log_manager._generate_content(
            sample_merged_df, sample_percentages, {}, [1001, 1002, 1003], []
        )

        assert "Missing SP: 1001, 1002, 1003" in content

    def test_generate_content_with_gun_timing(self, line_log_manager, sample_merged_df,
                                             sample_percentages):
        """Test content generation with gun timing data."""
        log_data = {
            'log_timing_warning': [(1001, ['G1', 'G2']), (1003, ['G3'])],
            'log_timing_error': [(1005, ['G4', 'G5'])],
        }

        content = line_log_manager._generate_content(
            sample_merged_df, sample_percentages, log_data, [], []
        )

        assert "1001 (G1,G2)" in content
        assert "1003 (G3)" in content
        assert "1005 (G4,G5)" in content

    def test_generate_content_with_repeatability_flag(self, line_log_manager, sample_merged_df,
                                                     sample_percentages):
        """Test content generation with repeatability flags (uses detect_range)."""
        log_data = {
            'log_repeatability_flag': [1001, 1003, 1005, 1007],
        }

        content = line_log_manager._generate_content(
            sample_merged_df, sample_percentages, log_data, [], []
        )

        # Should use detect_range for repeatability_flag
        assert "Total 4 SP" in content
        assert "1001-1007" in content

    def test_generate_content_excludes_gun_timing_flag(self, line_log_manager, sample_merged_df,
                                                       sample_percentages):
        """Test that log_gun_timing_flag is excluded from output."""
        log_data = {
            'log_gun_timing_flag': [1001, 1002, 1003],
            'log_gun_depth_flag': [1005],
        }

        content = line_log_manager._generate_content(
            sample_merged_df, sample_percentages, log_data, [], []
        )

        # log_gun_timing_flag should not appear
        assert "log_gun_timing_flag" not in content
        # But gun_depth_flag should appear
        assert "Gun Depth" in content

    def test_generate_content_with_sensor_violations(self, line_log_manager, sample_merged_df,
                                                    sample_percentages):
        """Test content generation with sensor violations."""
        log_data = {
            'log_gun_depth_sensor_violation': ['Sensor 1: 1001-1003', 'Sensor 2: 1005-1007'],
        }

        content = line_log_manager._generate_content(
            sample_merged_df, sample_percentages, log_data, [], []
        )

        assert "Sensor 1: 1001-1003" in content
        assert "Sensor 2: 1005-1007" in content

    def test_generate_content_with_consecutive_errors(self, line_log_manager, sample_merged_df,
                                                     sample_percentages):
        """Test content generation with consecutive error windows."""
        log_data = {
            'log_consec_7_source_errors': ['1010-1024'],
            'log_window_12_of_24_source_errors': ['1030-1054'],
            'log_window_16_of_40_source_errors': ['1060-1100'],
        }

        content = line_log_manager._generate_content(
            sample_merged_df, sample_percentages, log_data, [], []
        )

        assert "1010-1024" in content
        assert "1030-1054" in content
        assert "1060-1100" in content


class TestGetLabelForKey:
    """Test _get_label_for_key method."""

    def test_get_label_for_known_keys(self, line_log_manager):
        """Test getting labels for known keys."""
        assert "Sub-Array Sep <6.8m" in line_log_manager._get_label_for_key('log_sub_array_sep_flag')
        assert "Gun Depth <6m" in line_log_manager._get_label_for_key('log_gun_depth_flag')
        assert "Gun Timing >1.0ms" in line_log_manager._get_label_for_key('log_timing_warning')
        assert "Gun Timing >1.5ms" in line_log_manager._get_label_for_key('log_timing_error')

    def test_get_label_for_enhanced_qc_checks(self, line_log_manager):
        """Test getting labels for enhanced QC check keys (Phase 4.3)."""
        assert "Sub-Array Separation Percentage Violation" in line_log_manager._get_label_for_key(
            'log_sub_array_sep_percent_violation'
        )
        assert "7+ Consecutive SP" in line_log_manager._get_label_for_key('log_consec_7_source_errors')
        assert "12+ Source Errors" in line_log_manager._get_label_for_key('log_window_12_of_24_source_errors')

    def test_get_label_for_unknown_key(self, line_log_manager):
        """Test getting label for unknown key (returns key itself)."""
        unknown_key = 'log_unknown_flag'
        assert line_log_manager._get_label_for_key(unknown_key) == unknown_key


class TestFindCommentsCell:
    """Test _find_comments_cell method."""

    def test_find_comments_cell_success(self, line_log_manager):
        """Test finding comments cell successfully."""
        wb = Workbook()
        sheet = wb.active

        # Set up test data
        sheet['B5'].value = 'Acquisition and Processing Comments'
        sheet['B6'].value = 'Existing content'

        result = line_log_manager._find_comments_cell(sheet)
        assert result is not None
        assert result.coordinate == 'B6'

        wb.close()

    def test_find_comments_cell_not_found(self, line_log_manager):
        """Test when comments cell is not found."""
        wb = Workbook()
        sheet = wb.active

        # Don't add the label
        result = line_log_manager._find_comments_cell(sheet)
        assert result is None

        wb.close()

    def test_find_comments_cell_with_custom_label(self, mock_config):
        """Test finding comments cell with custom label."""
        mock_config.set('LineLog', 'acquisition_comments_label', 'Custom Comments Label')
        manager = LineLogManager(mock_config)

        wb = Workbook()
        sheet = wb.active
        sheet['B10'].value = 'Custom Comments Label'

        result = manager._find_comments_cell(sheet)
        assert result is not None
        assert result.coordinate == 'B11'

        wb.close()


class TestDetectRange:
    """Test detect_range static method."""

    def test_detect_range_single_point(self, line_log_manager):
        """Test range detection with single point."""
        result = line_log_manager.detect_range([1001])
        assert result == "Total 1 SP. 1001"

    def test_detect_range_consecutive_2_step(self, line_log_manager):
        """Test range detection with consecutive 2-step intervals."""
        result = line_log_manager.detect_range([1001, 1003, 1005, 1007])
        assert result == "Total 4 SP. 1001-1007"

    def test_detect_range_multiple_ranges(self, line_log_manager):
        """Test range detection with multiple ranges."""
        result = line_log_manager.detect_range([1001, 1003, 1005, 1011, 1013, 1015, 1017, 1031])
        assert result == "Total 8 SP. 1001-1005, 1011-1017, 1031"

    def test_detect_range_unsorted_input(self, line_log_manager):
        """Test range detection with unsorted input (should sort automatically)."""
        result = line_log_manager.detect_range([1007, 1001, 1005, 1003])
        assert result == "Total 4 SP. 1001-1007"

    def test_detect_range_empty_list(self, line_log_manager):
        """Test range detection with empty list."""
        result = line_log_manager.detect_range([])
        assert result == ""

    def test_detect_range_non_consecutive(self, line_log_manager):
        """Test range detection with non-consecutive points."""
        result = line_log_manager.detect_range([1001, 1010, 1020])
        assert result == "Total 3 SP. 1001, 1010, 1020"

    def test_detect_range_mixed_consecutive_and_isolated(self, line_log_manager):
        """Test range detection with mix of consecutive and isolated points."""
        result = line_log_manager.detect_range([1001, 1003, 1005, 1020, 1030, 1032])
        assert result == "Total 6 SP. 1001-1005, 1020, 1030-1032"


class TestUpdateLineLog:
    """Test update_line_log method (integration-style tests)."""

    def test_update_line_log_file_not_found(self, line_log_manager, sample_merged_df,
                                           sample_percentages, sample_log_data):
        """Test update when line log file doesn't exist."""
        result = line_log_manager.update_line_log(
            '/nonexistent/file.xlsm',
            sample_merged_df,
            sample_log_data,
            [],
            sample_percentages,
            []
        )
        assert result is False

    def test_update_line_log_success(self, line_log_manager, sample_merged_df,
                                    sample_percentages, sample_log_data, tmp_path):
        """Test successful line log update."""
        # Create a test Excel file
        wb = Workbook()
        sheet = wb.active
        sheet['B5'].value = 'Acquisition and Processing Comments'
        sheet['B6'].value = 'Old content'
        sheet['E6'].value = 'Old date'

        test_file = tmp_path / "test_linelog.xlsm"
        wb.save(str(test_file))
        wb.close()

        # Update the line log
        result = line_log_manager.update_line_log(
            str(test_file),
            sample_merged_df,
            sample_log_data,
            [],
            sample_percentages,
            []
        )

        assert result is True

        # Verify the update
        wb = Workbook()
        wb = line_log_manager.open_workbook_with_retry(str(test_file))
        sheet = wb.active

        # Check date was updated
        assert sheet['E6'].value == '01-Oct-25'

        # Check content was updated
        content = sheet['B6'].value
        assert "Shooting Mode: 4D Source" in content
        assert content != 'Old content'

        wb.close()

    def test_update_line_log_without_datetime_column(self, line_log_manager,
                                                    sample_percentages, sample_log_data, tmp_path):
        """Test update when DataFrame doesn't have datetime_UTC column."""
        # Create DataFrame without datetime_UTC
        df = pd.DataFrame({
            'shot_point': [1001, 1002, 1003]
        })

        # Create test Excel file
        wb = Workbook()
        sheet = wb.active
        sheet['B5'].value = 'Acquisition and Processing Comments'
        sheet['B6'].value = 'Old content'
        sheet['E6'].value = 'Old date'

        test_file = tmp_path / "test_linelog_no_datetime.xlsm"
        wb.save(str(test_file))
        wb.close()

        # Update should still succeed
        result = line_log_manager.update_line_log(
            str(test_file),
            df,
            sample_log_data,
            [],
            sample_percentages,
            []
        )

        assert result is True

        # Verify date was NOT updated (remains old value)
        wb = line_log_manager.open_workbook_with_retry(str(test_file))
        sheet = wb.active
        assert sheet['E6'].value == 'Old date'

        wb.close()

    def test_update_line_log_missing_comments_cell(self, line_log_manager, sample_merged_df,
                                                   sample_percentages, sample_log_data, tmp_path):
        """Test update when comments cell is not found."""
        # Create Excel file WITHOUT the comments label
        wb = Workbook()
        sheet = wb.active
        sheet['B5'].value = 'Wrong Label'

        test_file = tmp_path / "test_linelog_no_label.xlsm"
        wb.save(str(test_file))
        wb.close()

        # Update should fail
        result = line_log_manager.update_line_log(
            str(test_file),
            sample_merged_df,
            sample_log_data,
            [],
            sample_percentages,
            []
        )

        assert result is False


class TestEdgeCases:
    """Test edge cases and error handling."""

    def test_generate_content_with_empty_dataframe(self, line_log_manager):
        """Test content generation with empty DataFrame."""
        empty_df = pd.DataFrame()

        content = line_log_manager._generate_content(
            empty_df, {'percent_radial': 0, 'percent_gd_errors': 0}, {}, [], []
        )

        assert "Shooting Mode: 4D Source" in content
        assert "100.00%" in content  # 100 - 0

    def test_generate_content_with_none_percentages(self, line_log_manager, sample_merged_df):
        """Test content generation when percentages dict doesn't have keys (defaults to 0)."""
        percentages = {}  # Empty dict, .get() will return 0

        # Should handle gracefully with .get() fallback
        content = line_log_manager._generate_content(
            sample_merged_df, percentages, {}, [], []
        )

        assert "100.00%" in content  # 100 - 0 = 100

    def test_detect_range_with_duplicates(self, line_log_manager):
        """Test range detection with duplicate shot points."""
        # sorted() preserves duplicates, so they create multiple ranges
        result = line_log_manager.detect_range([1001, 1003, 1003, 1005])
        # Expected: duplicate 1003 breaks the range
        assert "Total 4 SP" in result
        assert "1001-1003" in result or "1003-1005" in result

    def test_find_line_log_file_with_permission_error(self, line_log_manager, tmp_path):
        """Test finding file when directory has permission issues (hard to simulate)."""
        # This test would require OS-specific permission manipulation
        # Skipping for now, but documenting the test case
        pass


class TestExtractShotPointMarkers:
    """Test suite for shot point marker extraction functionality."""

    @pytest.fixture
    def create_test_workbook(self, tmp_path):
        """Create a test workbook with shot point markers."""
        def _create(markers_data):
            wb = Workbook()
            sheet = wb.active

            # Add data to the workbook based on markers_data
            # markers_data format: [(row, col_b, col_c, col_f), ...]
            for row, time_val, sp_val, marker_val in markers_data:
                sheet[f'B{row}'] = time_val
                sheet[f'C{row}'] = sp_val
                sheet[f'F{row}'] = marker_val

            # Save to temporary file
            file_path = tmp_path / "test_linelog.xlsm"
            wb.save(file_path)
            wb.close()

            return str(file_path)

        return _create

    def test_extract_all_markers(self, line_log_manager, create_test_workbook):
        """Test extraction of all marker types."""
        markers_data = [
            (18, '08:34:00', 6735, 'FASP'),
            (19, '08:37:00', 6803, 'FOSP'),
            (20, '08:38:00', 6821, 'LOSP'),
            (21, '08:39:00', 6823, 'FGSP. SOL.'),
            (23, '09:34:00', 7871, 'LGSP. EOL'),
            (25, '10:00:00', 8000, 'LSP'),
        ]

        file_path = create_test_workbook(markers_data)
        result = line_log_manager.extract_shot_point_markers(file_path)

        # Verify all markers found
        assert result['FASP'] is not None
        assert result['FASP']['sp'] == 6735
        assert result['FASP']['time'] == '08:34:00'
        assert result['FASP']['row'] == 18

        assert result['FGSP'] is not None
        assert result['FGSP']['sp'] == 6823
        assert result['FGSP']['description'] == 'FGSP. SOL.'

        assert result['LGSP'] is not None
        assert result['LGSP']['sp'] == 7871

        assert result['FOSP'] is not None
        assert result['FOSP']['sp'] == 6803

        assert result['LOSP'] is not None
        assert result['LOSP']['sp'] == 6821

        assert result['LSP'] is not None
        assert result['LSP']['sp'] == 8000

    def test_extract_markers_with_contaminated_text(self, line_log_manager, create_test_workbook):
        """Test extraction with real-world contaminated descriptions."""
        markers_data = [
            (18, '08:34:00', 6735, 'FASP'),
            (21, '08:39:00', 6823, 'FGSP. SOL. Continuation of Seq. 1886'),
            (23, '09:34:00', 7871, 'LGSP. Terminated earlier d/t no permission from P-38'),
        ]

        file_path = create_test_workbook(markers_data)
        result = line_log_manager.extract_shot_point_markers(file_path)

        # Verify markers found despite contaminated text
        assert result['FASP']['sp'] == 6735
        assert result['FGSP']['sp'] == 6823
        assert 'Continuation of Seq. 1886' in result['FGSP']['description']
        assert result['LGSP']['sp'] == 7871
        assert 'no permission' in result['LGSP']['description']

    def test_extract_markers_no_overlap(self, line_log_manager, create_test_workbook):
        """Test extraction when no overlap markers present."""
        markers_data = [
            (18, '08:34:00', 6735, 'FASP'),
            (21, '08:39:00', 6823, 'FGSP. SOL.'),
            (23, '09:34:00', 7871, 'LGSP. EOL'),
        ]

        file_path = create_test_workbook(markers_data)
        result = line_log_manager.extract_shot_point_markers(file_path)

        # Verify overlap markers are None
        assert result['FOSP'] is None
        assert result['LOSP'] is None
        assert result['LSP'] is None

        # Verify standard markers found
        assert result['FASP'] is not None
        assert result['FGSP'] is not None
        assert result['LGSP'] is not None

    def test_extract_markers_empty_cells(self, line_log_manager, create_test_workbook):
        """Test extraction with some empty cells in range."""
        markers_data = [
            (18, '08:34:00', 6735, 'FASP'),
            (19, None, None, None),  # Empty row
            (20, None, None, None),  # Empty row
            (21, '08:39:00', 6823, 'FGSP'),
        ]

        file_path = create_test_workbook(markers_data)
        result = line_log_manager.extract_shot_point_markers(file_path)

        assert result['FASP']['sp'] == 6735
        assert result['FGSP']['sp'] == 6823
        assert result['LGSP'] is None

    def test_extract_markers_missing_sp_column(self, line_log_manager, create_test_workbook):
        """Test extraction when shot point column has invalid data."""
        markers_data = [
            (18, '08:34:00', 'invalid', 'FASP'),  # Invalid SP
            (21, '08:39:00', 6823, 'FGSP'),
        ]

        file_path = create_test_workbook(markers_data)
        result = line_log_manager.extract_shot_point_markers(file_path)

        # FASP should be found but with None SP
        assert result['FASP'] is not None
        assert result['FASP']['sp'] is None
        assert result['FASP']['time'] == '08:34:00'

        # FGSP should be normal
        assert result['FGSP']['sp'] == 6823

    def test_extract_markers_custom_range(self, line_log_manager, create_test_workbook):
        """Test extraction with custom search range."""
        markers_data = [
            (10, '08:34:00', 6735, 'FASP'),  # Outside default range
            (30, '08:39:00', 6823, 'FGSP'),
        ]

        file_path = create_test_workbook(markers_data)

        # Test with custom range (10, 35)
        result = line_log_manager.extract_shot_point_markers(file_path, search_range=(10, 35))

        assert result['FASP']['sp'] == 6735
        assert result['FGSP']['sp'] == 6823

        # Test with default range (18, 50) - should not find FASP at row 10
        result_default = line_log_manager.extract_shot_point_markers(file_path)
        assert result_default['FASP'] is None
        assert result_default['FGSP']['sp'] == 6823

    def test_extract_markers_case_insensitive(self, line_log_manager, create_test_workbook):
        """Test that marker detection is case insensitive."""
        markers_data = [
            (18, '08:34:00', 6735, 'fasp'),  # lowercase
            (21, '08:39:00', 6823, 'FgSp'),  # mixed case
            (23, '09:34:00', 7871, 'LGSP'),  # uppercase
        ]

        file_path = create_test_workbook(markers_data)
        result = line_log_manager.extract_shot_point_markers(file_path)

        assert result['FASP']['sp'] == 6735
        assert result['FGSP']['sp'] == 6823
        assert result['LGSP']['sp'] == 7871

    def test_extract_markers_file_not_found(self, line_log_manager):
        """Test extraction when file doesn't exist."""
        result = line_log_manager.extract_shot_point_markers('/nonexistent/path.xlsm')

        # Should return all None markers
        assert result['FASP'] is None
        assert result['FGSP'] is None
        assert result['LGSP'] is None
        assert result['LSP'] is None
        assert result['FOSP'] is None
        assert result['LOSP'] is None

    def test_extract_markers_multiple_keywords_same_cell(self, line_log_manager, create_test_workbook):
        """Test when multiple keywords appear in same cell (should match first)."""
        markers_data = [
            (18, '08:34:00', 6735, 'FASP and FGSP combined'),  # Multiple keywords
        ]

        file_path = create_test_workbook(markers_data)
        result = line_log_manager.extract_shot_point_markers(file_path)

        # Should match FASP first (order in markers.keys())
        assert result['FASP'] is not None
        assert result['FASP']['sp'] == 6735
        # FGSP should not be matched from the same cell
        assert result['FGSP'] is None

    def test_extract_markers_has_overlap_check(self, line_log_manager, create_test_workbook):
        """Test helper logic for checking overlap scenario."""
        # With overlap
        markers_data = [
            (18, '08:34:00', 6735, 'FASP'),
            (19, '08:37:00', 6803, 'FOSP'),
            (20, '08:38:00', 6821, 'LOSP'),
            (21, '08:39:00', 6823, 'FGSP'),
        ]

        file_path = create_test_workbook(markers_data)
        result = line_log_manager.extract_shot_point_markers(file_path)

        has_overlap = result['FOSP'] is not None and result['LOSP'] is not None
        assert has_overlap is True

        # Without overlap
        markers_data_no_overlap = [
            (18, '08:34:00', 6735, 'FASP'),
            (21, '08:39:00', 6823, 'FGSP'),
        ]

        file_path2 = create_test_workbook(markers_data_no_overlap)
        result2 = line_log_manager.extract_shot_point_markers(file_path2)

        has_overlap2 = result2['FOSP'] is not None and result2['LOSP'] is not None
        assert has_overlap2 is False


class TestGetCellValue:
    """Test suite for _get_cell_value method."""

    @pytest.fixture
    def create_test_sheet(self):
        """Create a test worksheet with various cell values."""
        wb = Workbook()
        sheet = wb.active

        # Set up test cells
        sheet['C6'] = '3184P31885'
        sheet['C7'] = 'Line 1'
        sheet['C8'] = 1885
        sheet['C9'] = 1
        sheet['E8'] = 45.5
        sheet['B10'] = None  # Empty cell
        sheet['B11'] = '  spaced text  '

        return sheet, wb

    def test_get_cell_value_string(self, line_log_manager, create_test_sheet):
        """Test getting string value from cell."""
        sheet, wb = create_test_sheet

        result = line_log_manager._get_cell_value(sheet, 'cell_filename', 'C6', str)
        assert result == '3184P31885'

        wb.close()

    def test_get_cell_value_int(self, line_log_manager, create_test_sheet):
        """Test getting integer value from cell."""
        sheet, wb = create_test_sheet

        result = line_log_manager._get_cell_value(sheet, 'cell_sequence', 'C8', int)
        assert result == 1885
        assert isinstance(result, int)

        wb.close()

    def test_get_cell_value_float(self, line_log_manager, create_test_sheet):
        """Test getting float value from cell."""
        sheet, wb = create_test_sheet

        result = line_log_manager._get_cell_value(sheet, 'cell_heading', 'E8', float)
        assert result == 45.5
        assert isinstance(result, float)

        wb.close()

    def test_get_cell_value_none(self, line_log_manager, create_test_sheet):
        """Test getting value from empty cell."""
        sheet, wb = create_test_sheet

        result = line_log_manager._get_cell_value(sheet, 'cell_empty', 'B10', str)
        assert result is None

        wb.close()

    def test_get_cell_value_with_spaces(self, line_log_manager, create_test_sheet):
        """Test that string values are trimmed."""
        sheet, wb = create_test_sheet

        result = line_log_manager._get_cell_value(sheet, 'cell_spaced', 'B11', str)
        assert result == 'spaced text'

        wb.close()

    def test_get_cell_value_invalid_conversion(self, line_log_manager, create_test_sheet):
        """Test invalid type conversion returns None."""
        sheet, wb = create_test_sheet

        # Try to convert string to int (should fail gracefully)
        result = line_log_manager._get_cell_value(sheet, 'cell_filename', 'C6', int)
        assert result is None

        wb.close()

    def test_get_cell_value_from_config(self, create_test_sheet):
        """Test that cell reference is read from config."""
        sheet, wb = create_test_sheet

        # Create config with custom cell reference
        config = ConfigParser()
        config.add_section('LineLog')
        config.set('LineLog', 'custom_cell', 'C7')

        manager = LineLogManager(config)
        result = manager._get_cell_value(sheet, 'custom_cell', 'C6', str)

        # Should use C7 from config, not default C6
        assert result == 'Line 1'

        wb.close()


class TestExtractLineInfo:
    """Test suite for extract_line_info method."""

    @pytest.fixture
    def create_full_test_workbook(self, tmp_path):
        """Create a comprehensive test workbook with all data."""
        def _create(markers_data, metadata):
            wb = Workbook()
            sheet = wb.active

            # Add marker data
            for row, time_val, sp_val, marker_val in markers_data:
                sheet[f'B{row}'] = time_val
                sheet[f'C{row}'] = sp_val
                sheet[f'F{row}'] = marker_val

            # Add metadata
            sheet['C6'] = metadata.get('filename', '3184P31885')
            sheet['C7'] = metadata.get('line', 'Line 1')
            sheet['C8'] = metadata.get('sequence', 1885)
            sheet['C9'] = metadata.get('attempt', 1)
            sheet['E8'] = metadata.get('heading', 45.5)

            file_path = tmp_path / "test_full_linelog.xlsm"
            wb.save(file_path)
            wb.close()

            return str(file_path)

        return _create

    def test_extract_line_info_complete(self, mock_config, create_full_test_workbook):
        """Test extraction of complete line information."""
        mock_config.set('LineLog', 'shot_increment', '2')
        manager = LineLogManager(mock_config)

        markers_data = [
            (18, '08:34:00', 6735, 'FASP'),
            (21, '08:39:00', 6823, 'FGSP'),
            (23, '09:34:00', 7871, 'LGSP'),
        ]

        metadata = {
            'filename': '3184P31885',
            'line': 'Line 1',
            'sequence': 1885,
            'attempt': 1,
            'heading': 45.5
        }

        file_path = create_full_test_workbook(markers_data, metadata)
        result = manager.extract_line_info(file_path)

        # Verify markers
        assert result['markers']['FASP']['sp'] == 6735
        assert result['markers']['FGSP']['sp'] == 6823
        assert result['markers']['LGSP']['sp'] == 7871

        # Verify metadata
        assert result['metadata']['filename'] == '3184P31885'
        assert result['metadata']['line'] == 'Line 1'
        assert result['metadata']['sequence'] == 1885
        assert result['metadata']['attempt'] == 1
        assert result['metadata']['heading'] == 45.5

        # Verify calculated values
        assert result['calculated']['shot_increment'] == 2
        expected_production_sp = (abs(7871 - 6823) // 2) + 1
        assert result['calculated']['production_sp'] == expected_production_sp
        assert result['calculated']['has_overlap'] is False
        assert result['calculated']['overlap_sp'] is None

    def test_extract_line_info_with_overlap(self, mock_config, create_full_test_workbook):
        """Test extraction with overlap markers present."""
        mock_config.set('LineLog', 'shot_increment', '2')
        manager = LineLogManager(mock_config)

        markers_data = [
            (18, '08:34:00', 6735, 'FASP'),
            (19, '08:37:00', 6803, 'FOSP'),
            (20, '08:38:00', 6821, 'LOSP'),
            (21, '08:39:00', 6823, 'FGSP'),
            (23, '09:34:00', 7871, 'LGSP'),
        ]

        metadata = {'filename': '3184P31885'}

        file_path = create_full_test_workbook(markers_data, metadata)
        result = manager.extract_line_info(file_path)

        # Verify overlap detection
        assert result['calculated']['has_overlap'] is True
        expected_overlap_sp = (abs(6821 - 6803) // 2) + 1
        assert result['calculated']['overlap_sp'] == expected_overlap_sp

    def test_extract_line_info_file_not_found(self, line_log_manager):
        """Test extraction when file doesn't exist."""
        result = line_log_manager.extract_line_info('/nonexistent/file.xlsm')

        # Should return initialized structure with None values
        # extract_shot_point_markers returns dict with None values, not empty dict
        assert result['markers']['FASP'] is None
        assert result['markers']['FGSP'] is None
        assert result['metadata']['filename'] is None
        assert result['calculated']['production_sp'] is None

    def test_extract_line_info_missing_markers(self, mock_config, create_full_test_workbook):
        """Test extraction when some markers are missing."""
        mock_config.set('LineLog', 'shot_increment', '2')
        manager = LineLogManager(mock_config)

        markers_data = [
            (18, '08:34:00', 6735, 'FASP'),
            # Missing FGSP and LGSP
        ]

        metadata = {'filename': '3184P31885'}

        file_path = create_full_test_workbook(markers_data, metadata)
        result = manager.extract_line_info(file_path)

        # Should handle missing markers gracefully
        assert result['calculated']['production_sp'] is None


class TestErrorHandling:
    """Test suite for error handling and edge cases."""

    def test_open_workbook_with_retry_permission_error(self, line_log_manager, tmp_path, monkeypatch):
        """Test retry logic when file is locked."""
        import openpyxl

        # Create a test file
        wb = Workbook()
        test_file = tmp_path / "locked.xlsm"
        wb.save(str(test_file))
        wb.close()

        # Mock openpyxl.load_workbook to raise PermissionError first 2 times
        call_count = {'count': 0}

        original_load = openpyxl.load_workbook

        def mock_load_workbook(*args, **kwargs):
            call_count['count'] += 1
            if call_count['count'] <= 2:
                raise PermissionError("File is locked")
            return original_load(*args, **kwargs)

        monkeypatch.setattr(openpyxl, 'load_workbook', mock_load_workbook)

        # Should succeed on third attempt
        result = line_log_manager.open_workbook_with_retry(str(test_file))
        assert result is not None
        assert call_count['count'] == 3

        result.close()

    def test_open_workbook_with_retry_max_attempts_exceeded(self, line_log_manager, tmp_path, monkeypatch):
        """Test that retry logic fails after max attempts."""
        import openpyxl

        # Create a test file
        wb = Workbook()
        test_file = tmp_path / "always_locked.xlsm"
        wb.save(str(test_file))
        wb.close()

        # Mock to always raise PermissionError
        def mock_load_workbook(*args, **kwargs):
            raise PermissionError("File is always locked")

        monkeypatch.setattr(openpyxl, 'load_workbook', mock_load_workbook)

        # Should fail after max_attempts
        result = line_log_manager.open_workbook_with_retry(str(test_file))
        assert result is None

    def test_open_workbook_with_retry_other_exception(self, line_log_manager, tmp_path, monkeypatch):
        """Test handling of non-PermissionError exceptions."""
        import openpyxl

        # Create a test file
        wb = Workbook()
        test_file = tmp_path / "corrupt.xlsm"
        wb.save(str(test_file))
        wb.close()

        # Mock to raise different exception
        def mock_load_workbook(*args, **kwargs):
            raise ValueError("Corrupt file")

        monkeypatch.setattr(openpyxl, 'load_workbook', mock_load_workbook)

        # Should fail immediately without retry
        result = line_log_manager.open_workbook_with_retry(str(test_file))
        assert result is None

    def test_update_line_log_save_permission_error(self, line_log_manager, sample_merged_df,
                                                   sample_percentages, sample_log_data, tmp_path, monkeypatch):
        """Test handling of PermissionError during save."""
        import openpyxl

        # Create test file
        wb = Workbook()
        sheet = wb.active
        sheet['B5'].value = 'Acquisition and Processing Comments'
        sheet['B6'].value = 'Old content'

        test_file = tmp_path / "readonly_linelog.xlsm"
        wb.save(str(test_file))
        wb.close()

        # Mock save method to raise PermissionError
        def mock_save(*args, **kwargs):
            raise PermissionError("Cannot save file")

        monkeypatch.setattr(openpyxl.Workbook, 'save', mock_save)

        # Should fail with False return value
        result = line_log_manager.update_line_log(
            str(test_file),
            sample_merged_df,
            sample_log_data,
            [],
            sample_percentages,
            []
        )

        assert result is False

    def test_update_line_log_save_other_exception(self, line_log_manager, sample_merged_df,
                                                  sample_percentages, sample_log_data, tmp_path, monkeypatch):
        """Test handling of other exceptions during save."""
        import openpyxl

        # Create test file
        wb = Workbook()
        sheet = wb.active
        sheet['B5'].value = 'Acquisition and Processing Comments'
        sheet['B6'].value = 'Old content'

        test_file = tmp_path / "error_linelog.xlsm"
        wb.save(str(test_file))
        wb.close()

        # Mock save method to raise different exception
        def mock_save(*args, **kwargs):
            raise IOError("Disk full")

        monkeypatch.setattr(openpyxl.Workbook, 'save', mock_save)

        # Should fail with False return value
        result = line_log_manager.update_line_log(
            str(test_file),
            sample_merged_df,
            sample_log_data,
            [],
            sample_percentages,
            []
        )

        assert result is False

    def test_generate_content_with_string_messages(self, line_log_manager, sample_merged_df,
                                                   sample_percentages):
        """Test content generation with string message log data."""
        log_data = {
            'log_sub_array_sep_percent_violation': '15.5% of SP violate threshold',
            'log_sub_array_sep_avg_violation': 'Sequence average 9.1m exceeds 8.8m',
            'log_percent_3_total_source_errors': '3.5% of total SP have source errors',
        }

        content = line_log_manager._generate_content(
            sample_merged_df, sample_percentages, log_data, [], []
        )

        assert '15.5% of SP violate threshold' in content
        assert 'Sequence average 9.1m exceeds 8.8m' in content
        assert '3.5% of total SP have source errors' in content

    def test_extract_line_info_exception_handling(self, line_log_manager, tmp_path, monkeypatch):
        """Test exception handling in extract_line_info."""
        # Create a test file
        wb = Workbook()
        sheet = wb.active
        sheet['C6'] = 'Test'

        test_file = tmp_path / "test_info.xlsm"
        wb.save(str(test_file))
        wb.close()

        # Mock _get_cell_value to raise exception
        def mock_get_cell_value(*args, **kwargs):
            raise RuntimeError("Error reading cell")

        monkeypatch.setattr(LineLogManager, '_get_cell_value', mock_get_cell_value)

        # Should handle exception and return default structure
        result = line_log_manager.extract_line_info(str(test_file))

        # Should return structure with None values
        assert result['metadata']['filename'] is None
        assert result['calculated']['production_sp'] is None


class TestFilterLogDataByRange:
    """Test suite for _filter_log_data_by_range method (production shot filtering)."""

    def test_filter_simple_sp_list(self, line_log_manager):
        """Test filtering simple list of shot points."""
        log_data = {
            'log_gun_depth_flag': [6800, 6820, 6840, 6860, 6880],  # approach + production + overlap
        }

        # Filter to production range 6825-6875
        filtered = line_log_manager._filter_log_data_by_range(log_data, 6825, 6875)

        # Only 6840, 6860 should remain
        assert filtered['log_gun_depth_flag'] == [6840, 6860]

    def test_filter_tuple_list(self, line_log_manager):
        """Test filtering list of tuples (sp, [guns])."""
        log_data = {
            'log_timing_warning': [
                (6800, ['G1', 'G2']),  # approach
                (6830, ['G3']),         # production
                (6850, ['G4']),         # production
                (6900, ['G5'])          # overlap
            ]
        }

        filtered = line_log_manager._filter_log_data_by_range(log_data, 6825, 6875)

        # Only production shots should remain
        assert filtered['log_timing_warning'] == [(6830, ['G3']), (6850, ['G4'])]

    def test_filter_string_messages(self, line_log_manager):
        """Test that string messages are kept unchanged."""
        log_data = {
            'log_sub_array_sep_percent_violation': '15.5% of SP violate threshold',
            'log_sub_array_sep_avg_violation': 'Sequence average 9.1m exceeds 8.8m',
        }

        filtered = line_log_manager._filter_log_data_by_range(log_data, 6825, 6875)

        # String messages should remain unchanged
        assert filtered == log_data

    def test_filter_range_strings(self, line_log_manager):
        """Test filtering range strings like '6810-6830'."""
        log_data = {
            'log_consec_7_source_errors': ['6810-6824', '6830-6850', '6870-6890']
        }

        filtered = line_log_manager._filter_log_data_by_range(log_data, 6825, 6875)

        # First range overlaps (trim to 6825-6824 = invalid, skip)
        # Second range fully inside (6830-6850)
        # Third range overlaps (trim to 6870-6875)
        assert '6830-6850' in filtered['log_consec_7_source_errors']
        assert '6870-6875' in filtered['log_consec_7_source_errors']

    def test_filter_sensor_violation_ranges(self, line_log_manager):
        """Test filtering sensor violations with prefix."""
        log_data = {
            'log_gun_depth_sensor_violation': [
                'Sensor 1: 6810-6820',  # before production
                'Sensor 2: 6830-6850',  # in production
                'Sensor 3: 6860-6890'   # overlaps end
            ]
        }

        filtered = line_log_manager._filter_log_data_by_range(log_data, 6825, 6875)

        # Should keep production ranges with prefix
        assert any('6830-6850' in item for item in filtered['log_gun_depth_sensor_violation'])
        assert any('6860-6875' in item for item in filtered['log_gun_depth_sensor_violation'])

    def test_filter_empty_log_data(self, line_log_manager):
        """Test filtering with empty log data."""
        log_data = {}

        filtered = line_log_manager._filter_log_data_by_range(log_data, 6825, 6875)

        assert filtered == {}

    def test_filter_none_markers(self, line_log_manager):
        """Test that filtering is skipped if markers are None."""
        log_data = {
            'log_gun_depth_flag': [6800, 6820, 6840, 6860, 6880],
        }

        # If fgsp or lgsp is None, should return original data
        filtered = line_log_manager._filter_log_data_by_range(log_data, None, 6875)
        assert filtered == log_data

        filtered = line_log_manager._filter_log_data_by_range(log_data, 6825, None)
        assert filtered == log_data

    def test_filter_descending_sequence(self, line_log_manager):
        """Test filtering with descending shot points (LGSP < FGSP)."""
        log_data = {
            'log_gun_depth_flag': [6900, 6880, 6860, 6840, 6820, 6800],
        }

        # FGSP=6875, LGSP=6825 (descending) - range is 6825-6875
        filtered = line_log_manager._filter_log_data_by_range(log_data, 6875, 6825)

        # Should keep 6860, 6840 (between 6825 and 6875, excluding 6880 which is > 6875)
        assert set(filtered['log_gun_depth_flag']) == {6860, 6840}

    def test_filter_removes_empty_lists(self, line_log_manager):
        """Test that empty filtered lists are removed from result."""
        log_data = {
            'log_gun_depth_flag': [6800, 6820],  # All in approach
            'log_volume_flag': [6840, 6860],      # All in production
        }

        filtered = line_log_manager._filter_log_data_by_range(log_data, 6825, 6875)

        # log_gun_depth_flag should be removed (empty after filter)
        assert 'log_gun_depth_flag' not in filtered
        # log_volume_flag should remain
        assert 'log_volume_flag' in filtered


class TestGenerateContentWithFiltering:
    """Test _generate_content with production shot filtering and overlap."""

    def test_generate_content_with_filtering(self, line_log_manager, sample_merged_df,
                                            sample_percentages):
        """Test content generation with production filtering."""
        log_data = {
            'log_gun_depth_flag': [6800, 6825, 6850, 6875, 6900],  # Mix of approach/production/overlap
            'log_repeatability_flag': [6810, 6830, 6850, 6870, 6890],
        }

        # Generate with FGSP=6820, LGSP=6880 (should filter to production only)
        content = line_log_manager._generate_content(
            sample_merged_df, sample_percentages, log_data, [], [],
            fgsp=6820, lgsp=6880, fosp=None, losp=None
        )

        # Should only contain production SPs (6825, 6850, 6875, 6830, 6870)
        assert '6825' in content or '6850' in content or '6875' in content
        # Should NOT contain approach/overlap SPs
        assert '6800' not in content
        assert '6900' not in content

    def test_generate_content_with_overlap_comment(self, line_log_manager, sample_merged_df,
                                                   sample_percentages):
        """Test that overlap comment is added when FOSP/LOSP provided."""
        log_data = {
            'log_gun_depth_flag': [6850],
        }

        content = line_log_manager._generate_content(
            sample_merged_df, sample_percentages, log_data, [], [],
            fgsp=6823, lgsp=7871, fosp=6803, losp=6821
        )

        # Should include overlap comment
        assert 'SP 6803-6821 overlap' in content

    def test_generate_content_without_overlap(self, line_log_manager, sample_merged_df,
                                             sample_percentages):
        """Test that no overlap comment is added when FOSP/LOSP are None."""
        log_data = {
            'log_gun_depth_flag': [6850],
        }

        content = line_log_manager._generate_content(
            sample_merged_df, sample_percentages, log_data, [], [],
            fgsp=6823, lgsp=7871, fosp=None, losp=None
        )

        # Should NOT include overlap comment
        assert 'overlap' not in content.lower()

    def test_generate_content_no_markers(self, line_log_manager, sample_merged_df,
                                        sample_percentages):
        """Test content generation without markers (backward compatibility)."""
        log_data = {
            'log_gun_depth_flag': [6800, 6825, 6850],
        }

        # No markers provided - should not filter
        content = line_log_manager._generate_content(
            sample_merged_df, sample_percentages, log_data, [], []
        )

        # All SPs should be present (no filtering)
        assert '6800' in content or '6825' in content or '6850' in content
