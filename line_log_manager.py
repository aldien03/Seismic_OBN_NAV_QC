"""
Line Log Manager Module

This module provides line log update functionality for Excel (.xlsm) files.
It handles finding, opening, updating, and saving line log workbooks with QC data.

Classes:
- LineLogManager: Main class for handling line log operations

Author: aldien03@gmail.com
Date: 2025-09-30
"""

import os
import re
import time
import logging
import pandas as pd
import openpyxl
from typing import Dict, List, Optional, Tuple
from configparser import ConfigParser
from openpyxl.styles import Alignment


class LineLogManager:
    """
    Class for managing Excel line log operations.

    Supports:
    - Finding line log files (.xlsm) in directory
    - Opening workbooks with retry logic for locked files
    - Updating date, FGSP/LGSP data
    - Formatting and inserting QC comments
    - Handling percentages, missing SP, and log data
    - Saving workbooks with proper error handling
    """

    def __init__(self, config: ConfigParser):
        """
        Initialize LineLogManager with configuration.

        Args:
            config: ConfigParser instance containing line log settings
        """
        self.config = config
        self.line_log_pattern = r'0256-\d{4}[A-Z]\d\d{4}_Nav_LineLog\.xlsm$'
        self.max_attempts = config.getint('LineLog', 'max_open_attempts', fallback=5)
        self.comments_label = config.get('LineLog', 'acquisition_comments_label',
                                        fallback='Acquisition and Processing Comments')

    def find_line_log_file(self, directory: str) -> Optional[str]:
        """
        Find line log file in directory.

        Args:
            directory: Directory to search in

        Returns:
            Full path to line log file, or None if not found
        """
        if not os.path.isdir(directory):
            logging.warning(f"Invalid directory: {directory}")
            return None

        line_log_files = [f for f in os.listdir(directory) if re.match(self.line_log_pattern, f)]
        logging.debug(f"Found {len(line_log_files)} Line Log files")

        if not line_log_files:
            logging.warning("Line Log file not found in directory")
            return None

        line_log_path = os.path.join(directory, line_log_files[0])
        logging.debug(f"Line Log path: {line_log_path}")
        return line_log_path

    def open_workbook_with_retry(self, file_path: str) -> Optional[openpyxl.Workbook]:
        """
        Open workbook with retry logic for locked files.

        Args:
            file_path: Path to Excel file

        Returns:
            Opened workbook, or None if failed
        """
        wb = None
        for attempt in range(self.max_attempts):
            try:
                wb = openpyxl.load_workbook(file_path, keep_vba=True)
                logging.debug(f"Successfully opened Line Log file on attempt {attempt + 1}")
                break
            except PermissionError:
                if attempt < self.max_attempts - 1:
                    logging.warning(f"Line Log file is locked, attempt {attempt + 1} of {self.max_attempts}")
                    time.sleep(2)
                else:
                    logging.error("Unable to open Line Log file after multiple attempts")
                    return None
            except Exception as e:
                logging.error(f"Failed to open Line Log: {str(e)}")
                return None

        return wb

    def update_line_log(self, file_path: str, merged_df: pd.DataFrame,
                       log_data: Dict, missed_sp: List, percentages: Dict,
                       consecutive_errors: List,
                       fgsp: int = None, lgsp: int = None,
                       fosp: int = None, losp: int = None) -> bool:
        """
        Update line log file with QC data.

        Filters log data to production shots only (FGSP to LGSP) and adds
        overlap comment if applicable.

        Args:
            file_path: Path to line log .xlsm file
            merged_df: Merged DataFrame containing QC results
            log_data: Dictionary of log data from QC validation
            missed_sp: List of missing shot points
            percentages: Dictionary of QC error percentages
            consecutive_errors: List of consecutive error ranges
            fgsp: First Good Shot Point (optional, for filtering)
            lgsp: Last Good Shot Point (optional, for filtering)
            fosp: First Overlap Shot Point (optional, for overlap comment)
            losp: Last Overlap Shot Point (optional, for overlap comment)

        Returns:
            True if successful, False otherwise
        """
        wb = self.open_workbook_with_retry(file_path)
        if not wb:
            return False

        sheet = wb.active

        try:
            logging.debug("Updating Line Log content")

            # Update date in cell E6 if datetime_UTC exists
            if 'datetime_UTC' in merged_df.columns and not merged_df.empty:
                first_date = merged_df['datetime_UTC'].iloc[0].strftime('%d-%b-%y')
                logging.debug(f"Updating date in E6 to: {first_date}")
                sheet['E6'].value = first_date

            # Generate content (with production shot filtering if markers provided)
            content = self._generate_content(merged_df, percentages, log_data, missed_sp,
                                            consecutive_errors, fgsp, lgsp, fosp, losp)

            # Find and update target cell
            target_cell = self._find_comments_cell(sheet)
            if target_cell:
                target_cell.value = content
                target_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                logging.error(f"Could not find '{self.comments_label}' in the Line Log")
                return False

            # Save the workbook
            try:
                wb.save(file_path)
                logging.debug("Successfully saved Line Log")
                return True
            except PermissionError:
                logging.error("PermissionError when saving Line Log")
                return False
            except Exception as e:
                logging.error(f"Failed to save Line Log: {str(e)}")
                return False

        except Exception as e:
            logging.error(f"Unexpected error in update_line_log: {str(e)}")
            return False
        finally:
            if wb:
                wb.close()
                logging.debug("Closed Line Log workbook")

    def _filter_log_data_by_range(self, log_data: Dict, fgsp: int, lgsp: int) -> Dict:
        """
        Filter log data to only include shot points between FGSP and LGSP (production shots).

        Filters out shot points in approach (before FGSP) and overlap (after LGSP if descending).
        Handles different log data types:
        - Simple lists: [sp1, sp2, sp3]
        - Tuple lists: [(sp, [guns]), ...]
        - String messages: "message text" (unchanged)
        - Range strings: ["1001-1005", ...] (filter and reconstruct)

        Args:
            log_data: Dictionary of log data with shot points
            fgsp: First Good Shot Point
            lgsp: Last Good Shot Point

        Returns:
            Filtered log data dictionary with only production shots
        """
        if not log_data or fgsp is None or lgsp is None:
            return log_data

        filtered_data = {}
        min_sp = min(fgsp, lgsp)
        max_sp = max(fgsp, lgsp)

        for key, value in log_data.items():
            if not value:
                continue

            # String messages - keep as is
            if isinstance(value, str):
                filtered_data[key] = value
                continue

            # List of tuples (sp, [guns])
            if isinstance(value, list) and value and isinstance(value[0], tuple):
                filtered_list = [(sp, guns) for sp, guns in value
                               if min_sp <= sp <= max_sp]
                if filtered_list:
                    filtered_data[key] = filtered_list

            # List of range strings like ['1001-1005', '1010-1020']
            elif key in ['log_consec_7_source_errors', 'log_window_12_of_24_source_errors',
                        'log_window_16_of_40_source_errors', 'log_gun_depth_sensor_violation']:
                # Parse ranges, filter, and reconstruct
                filtered_ranges = []
                for item in value:
                    if '-' in str(item):
                        # Parse range string
                        parts = str(item).split('-')
                        if len(parts) == 2:
                            try:
                                range_start = int(parts[0].strip().split()[-1])  # Handle "Sensor 1: 1001"
                                range_end = int(parts[1].strip().split()[0])     # Handle "1005 more text"

                                # Check if range overlaps with production range
                                if range_end >= min_sp and range_start <= max_sp:
                                    # Trim range to production bounds
                                    trimmed_start = max(range_start, min_sp)
                                    trimmed_end = min(range_end, max_sp)

                                    # Reconstruct range string with original prefix (e.g., "Sensor 1: ")
                                    prefix = item.split(str(range_start))[0] if str(range_start) in item else ""
                                    filtered_ranges.append(f"{prefix}{trimmed_start}-{trimmed_end}")
                            except (ValueError, IndexError):
                                # Keep unparseable items as-is
                                filtered_ranges.append(item)
                        else:
                            filtered_ranges.append(item)
                    else:
                        # Single SP
                        try:
                            sp = int(item)
                            if min_sp <= sp <= max_sp:
                                filtered_ranges.append(item)
                        except ValueError:
                            filtered_ranges.append(item)

                if filtered_ranges:
                    filtered_data[key] = filtered_ranges

            # Simple list of shot points
            elif isinstance(value, list):
                filtered_list = [sp for sp in value if min_sp <= sp <= max_sp]
                if filtered_list:
                    filtered_data[key] = filtered_list

            else:
                # Unknown type, keep as is
                filtered_data[key] = value

        return filtered_data

    def _generate_content(self, merged_df: pd.DataFrame, percentages: Dict,
                         log_data: Dict, missed_sp: List, consecutive_errors: List,
                         fgsp: int = None, lgsp: int = None,
                         fosp: int = None, losp: int = None) -> str:
        """
        Generate content string for line log comments.

        Filters log data to only include production shots (FGSP to LGSP).
        Adds overlap comment if FOSP and LOSP are present.

        Args:
            merged_df: Merged DataFrame
            percentages: Percentages dictionary
            log_data: Log data dictionary
            missed_sp: List of missing shot points
            consecutive_errors: Consecutive errors list
            fgsp: First Good Shot Point (optional)
            lgsp: Last Good Shot Point (optional)
            fosp: First Overlap Shot Point (optional)
            losp: Last Overlap Shot Point (optional)

        Returns:
            Formatted content string with production shots only and overlap info
        """
        content = f"""Shooting Mode: 4D Source
Percentage of shotpoints with center of source at or within 10m radial distance from preplot = {100 - percentages.get('percent_radial', 0):.2f}%
Percentage of shotpoints with Average depth of active source array at or within 1m from nominal 7m depth = {100 - percentages.get('percent_gd_errors', 0):.2f}%"""

        # Filter log data to production shots only (FGSP to LGSP)
        if fgsp is not None and lgsp is not None:
            log_data = self._filter_log_data_by_range(log_data, fgsp, lgsp)
            logging.info("Filtered log_data to production range FGSP=%s to LGSP=%s", fgsp, lgsp)

        logging.info("log_autofires in log_data: %s", log_data.get('log_autofires', []))

        # Add additional information if available
        additional_info = []
        logging.info("log_data: %s", log_data)

        if missed_sp:
            additional_info.append(f"Missing SP: {', '.join(map(str, missed_sp))}")

        # List of keys to exclude from the output
        excluded_keys = ['log_gun_timing_flag']

        for key, value in log_data.items():
            # Skip excluded keys
            if key in excluded_keys:
                continue

            if value:
                label = self._get_label_for_key(key)

                # Format based on value type
                if key in ['log_timing_warning','log_timing_error','log_gun_disabled_flag','log_misfire_flag']:
                    # Gun-specific entries with tuple format (sp, [guns])
                    formatted_values = [f"{sp} ({','.join(guns)})" for sp, guns in value]
                    additional_info.append(f"{label}: {', '.join(formatted_values)}")
                elif key == 'log_repeatability_flag':
                    # Use range detection for repeatability flag
                    range_summary = self.detect_range(value)
                    additional_info.append(f"{label}: {range_summary}")
                elif key in ['log_sub_array_sep_percent_violation', 'log_sub_array_sep_avg_violation',
                            'log_percent_3_total_source_errors']:
                    # String messages (no further formatting needed)
                    additional_info.append(f"{label}: {value}")
                elif key == 'log_gun_depth_sensor_violation':
                    # List of sensor warning strings
                    additional_info.append(f"{label}: {', '.join(value)}")
                elif key in ['log_consec_7_source_errors', 'log_window_12_of_24_source_errors',
                            'log_window_16_of_40_source_errors']:
                    # List of range strings already formatted
                    additional_info.append(f"{label}: {', '.join(value)}")
                else:
                    # Default: list of shot points
                    formatted_values = map(str, value)
                    additional_info.append(f"{label}: {', '.join(formatted_values)}")

        logging.info("additional_info: %s", additional_info)

        if additional_info:
            content += "\n" + "\n".join(additional_info)

        # Add overlap comment if FOSP and LOSP are present
        if fosp is not None and losp is not None:
            content += f"\nSP {fosp}-{losp} overlap"
            logging.info("Added overlap comment: SP %s-%s", fosp, losp)

        return content

    def _get_label_for_key(self, key: str) -> str:
        """
        Get human-readable label for log data key.

        Args:
            key: Log data key

        Returns:
            Human-readable label
        """
        label_map = {
            # Original QC checks
            'log_sub_array_sep_flag': "SP with Sub-Array Sep <6.8m or >9.2m",
            'log_gun_depth_flag': "SP with Gun Depth <6m or >8m",
            'log_volume_flag': "SP with Volume <3040 cui",
            'log_timing_warning': "SP with Gun Timing >1.0ms and <=1.5ms",
            'log_timing_error': "SP with Gun Timing >1.5ms",
            'log_gun_pressure_flag': "SP with Pressure <1900psi or >2100psi",
            'log_sma_flag': "SP with SMA >3",
            'log_gun_disabled_flag': "SP with Gun Disabled",
            'log_misfire_flag': "SP with suspected misfire",
            'log_autofires': "SP with suspected autofire",
            'log_repeatability_flag': "SP with Radial >10.0m",

            # Enhanced QC checks (Phase 4.3)
            'log_sub_array_sep_percent_violation': "Sub-Array Separation Percentage Violation",
            'log_sub_array_sep_avg_violation': "Sub-Array Separation Sequence Average Violation",
            'log_gun_depth_sensor_violation': "Gun Depth Sensor Violations",
            'log_consec_7_source_errors': "7+ Consecutive SP with Source Errors",
            'log_window_12_of_24_source_errors': "12+ Source Errors in 24 SP Window",
            'log_window_16_of_40_source_errors': "16+ Source Errors in 40 SP Window",
            'log_percent_3_total_source_errors': "Source Errors Exceed 3% of Total SP",
        }
        return label_map.get(key, key)

    def _find_comments_cell(self, sheet) -> Optional[openpyxl.cell.Cell]:
        """
        Find the cell with "Acquisition and Processing Comments" label.

        Args:
            sheet: Excel worksheet

        Returns:
            Target cell (one row below label), or None if not found
        """
        for row in sheet['B:B']:
            if row.value == self.comments_label:
                target_cell = sheet.cell(row=row.row + 1, column=row.column)
                return target_cell
        return None

    def _get_cell_value(self, sheet, config_key: str, default_cell: str,
                       value_type: type = str):
        """
        Get cell value from Excel sheet using configurable cell reference.

        Args:
            sheet: Excel worksheet object
            config_key: Configuration key for cell reference (e.g., 'cell_filename')
            default_cell: Default cell reference if config key not found
            value_type: Type to convert value to (str, int, float)

        Returns:
            Parsed value from cell, or None if cell is empty or parsing fails
        """
        cell_ref = self.config.get('LineLog', config_key, fallback=default_cell)
        cell_value = sheet[cell_ref].value

        if cell_value is None:
            return None

        try:
            if value_type == str:
                return str(cell_value).strip()
            elif value_type == int:
                return int(cell_value)
            elif value_type == float:
                return float(cell_value)
            else:
                return cell_value
        except (ValueError, TypeError) as e:
            logging.warning(f"Could not parse cell {cell_ref} as {value_type.__name__}: {cell_value}")
            return None

    def extract_line_info(self, file_path: str) -> Dict:
        """
        Extract comprehensive line log information including markers, metadata, and calculations.

        Extracts:
        - Shot point markers (FASP, FGSP, LGSP, LSP, FOSP, LOSP)
        - Line metadata (filename, line, sequence, attempt, heading)
        - Calculated values (production SP, overlap SP, shot increment)

        Args:
            file_path: Path to line log .xlsm file

        Returns:
            Dictionary with complete line information:
            {
                'markers': {
                    'FASP': {'time': str, 'sp': int, 'row': int, 'description': str},
                    'FGSP': {...},
                    'LGSP': {...},
                    'LSP': {...} or None,
                    'FOSP': {...} or None,
                    'LOSP': {...} or None
                },
                'metadata': {
                    'filename': str,      # From cell C6
                    'line': str,          # From cell C7
                    'sequence': int,      # From cell C8
                    'attempt': int,       # From cell C9
                    'heading': float      # From cell E8
                },
                'calculated': {
                    'production_sp': int,      # ABS(LGSP-FGSP)/shot_increment + 1
                    'shot_increment': int,     # From config.ini
                    'has_overlap': bool,       # FOSP and LOSP both present
                    'overlap_sp': int or None  # ABS(LOSP-FOSP)/shot_increment + 1
                }
            }

        Example:
            line_info = manager.extract_line_info('path/to/linelog.xlsm')
            production_sp = line_info['calculated']['production_sp']
            line_name = line_info['metadata']['line']
            fgsp = line_info['markers']['FGSP']['sp']
        """
        # Initialize result structure
        result = {
            'markers': {},
            'metadata': {
                'filename': None,
                'line': None,
                'sequence': None,
                'attempt': None,
                'heading': None
            },
            'calculated': {
                'production_sp': None,
                'shot_increment': None,
                'has_overlap': False,
                'overlap_sp': None
            }
        }

        # Extract markers using existing method
        result['markers'] = self.extract_shot_point_markers(file_path)

        # Open workbook for metadata extraction
        wb = self.open_workbook_with_retry(file_path)
        if not wb:
            logging.error("Failed to open workbook for line info extraction")
            return result

        try:
            sheet = wb.active

            # Extract metadata from configured cells
            result['metadata']['filename'] = self._get_cell_value(sheet, 'cell_filename', 'C6', str)
            result['metadata']['line'] = self._get_cell_value(sheet, 'cell_line', 'C7', str)
            result['metadata']['sequence'] = self._get_cell_value(sheet, 'cell_sequence', 'C8', int)
            result['metadata']['attempt'] = self._get_cell_value(sheet, 'cell_attempt', 'C9', int)
            result['metadata']['heading'] = self._get_cell_value(sheet, 'cell_heading', 'E8', float)

            # Get shot increment from config
            shot_increment = self.config.getint('LineLog', 'shot_increment', fallback=2)
            result['calculated']['shot_increment'] = shot_increment

            # Calculate production SP
            if result['markers']['FGSP'] and result['markers']['LGSP']:
                fgsp = result['markers']['FGSP']['sp']
                lgsp = result['markers']['LGSP']['sp']
                if fgsp is not None and lgsp is not None:
                    production_sp = (abs(lgsp - fgsp) // shot_increment) + 1
                    result['calculated']['production_sp'] = production_sp
                    logging.debug(f"Calculated production SP: {production_sp} (FGSP={fgsp}, LGSP={lgsp}, increment={shot_increment})")

            # Check for overlap and calculate overlap SP
            has_overlap = (result['markers']['FOSP'] is not None and
                          result['markers']['LOSP'] is not None)
            result['calculated']['has_overlap'] = has_overlap

            if has_overlap:
                fosp = result['markers']['FOSP']['sp']
                losp = result['markers']['LOSP']['sp']
                if fosp is not None and losp is not None:
                    overlap_sp = (abs(losp - fosp) // shot_increment) + 1
                    result['calculated']['overlap_sp'] = overlap_sp
                    logging.debug(f"Calculated overlap SP: {overlap_sp} (FOSP={fosp}, LOSP={losp})")

            return result

        except Exception as e:
            logging.error(f"Error extracting line info: {str(e)}")
            return result
        finally:
            if wb:
                wb.close()

    def extract_shot_point_markers(self, file_path: str,
                                   search_column: str = None,
                                   search_range: Tuple[int, int] = None) -> Dict:
        """
        Extract shot point markers (FGSP, LGSP, FASP, LSP, FOSP, LOSP) from line log.

        Searches column F (rows 18-50 by default) for marker keywords and extracts:
        - Time (UTC) from column B
        - Shot Point number from column C
        - Full description from column F

        Args:
            file_path: Path to line log .xlsm file
            search_column: Column to search for markers (default from config or 'F')
            search_range: Tuple of (start_row, end_row) to search (default from config or (18, 50))

        Returns:
            Dictionary with marker data:
            {
                'FASP': {'time': str, 'sp': int, 'row': int, 'description': str},
                'FGSP': {...},
                'LGSP': {...},
                'LSP': {...} or None,
                'FOSP': {...} or None,
                'LOSP': {...} or None
            }

        Example:
            markers = manager.extract_shot_point_markers('path/to/linelog.xlsm')
            fgsp = markers['FGSP']['sp']  # 6823
            fgsp_time = markers['FGSP']['time']  # '08:39:00'
        """
        # Initialize all markers as None
        markers = {
            'FASP': None,
            'FGSP': None,
            'LGSP': None,
            'LSP': None,
            'FOSP': None,
            'LOSP': None
        }

        # Get search parameters from config if not provided
        if search_column is None:
            search_column = self.config.get('LineLog', 'marker_search_column', fallback='F')

        if search_range is None:
            start_row = self.config.getint('LineLog', 'marker_search_start_row', fallback=18)
            end_row = self.config.getint('LineLog', 'marker_search_end_row', fallback=50)
            search_range = (start_row, end_row)

        wb = self.open_workbook_with_retry(file_path)
        if not wb:
            logging.error("Failed to open workbook for marker extraction")
            return markers

        try:
            sheet = wb.active
            start_row, end_row = search_range

            for row_num in range(start_row, end_row + 1):
                cell_value = sheet[f'{search_column}{row_num}'].value

                if not cell_value:
                    continue

                cell_str = str(cell_value).strip()

                # Check for each marker keyword
                for marker_key in markers.keys():
                    if marker_key in cell_str.upper():
                        # Extract time from column B
                        time_cell = sheet[f'B{row_num}'].value
                        time_str = str(time_cell) if time_cell else None

                        # Extract shot point from column C
                        sp_cell = sheet[f'C{row_num}'].value
                        sp_num = None
                        if sp_cell is not None:
                            try:
                                sp_num = int(sp_cell)
                            except (ValueError, TypeError):
                                logging.warning(f"Could not parse SP at row {row_num}: {sp_cell}")

                        # Store marker data
                        markers[marker_key] = {
                            'time': time_str,
                            'sp': sp_num,
                            'row': row_num,
                            'description': cell_str
                        }

                        logging.debug(f"Found {marker_key} at row {row_num}: SP={sp_num}, Time={time_str}")
                        break  # Only match first marker keyword per cell

            return markers

        except Exception as e:
            logging.error(f"Error extracting shot point markers: {str(e)}")
            return markers
        finally:
            if wb:
                wb.close()

    def update_fasp_in_linelog(self, file_path: str, fasp_row: int, correct_sp: int, correct_time: str) -> bool:
        """
        Update FASP (First Actual Shot Point) in the line log Excel file.

        Args:
            file_path: Path to the line log .xlsm file
            fasp_row: Row number where FASP is located
            correct_sp: Correct shot point number to update
            correct_time: Correct time (HH:MM format) to update

        Returns:
            True if update successful, False otherwise
        """
        wb = None
        try:
            # Open workbook
            wb = self.open_workbook_with_retry(file_path)
            if not wb:
                logging.error(f"Could not open workbook: {file_path}")
                return False

            # Get the first sheet
            sheet = wb.worksheets[0]

            # Update Shot Point (Column C - typically column 3)
            sp_cell = sheet.cell(row=fasp_row, column=3)
            sp_cell.value = correct_sp
            logging.info(f"Updated FASP SP in row {fasp_row}, column C: {correct_sp}")

            # Update Time (Column B - typically column 2)
            time_cell = sheet.cell(row=fasp_row, column=2)
            time_cell.value = correct_time
            logging.info(f"Updated FASP Time in row {fasp_row}, column B: {correct_time}")

            # Save the workbook
            wb.save(file_path)
            logging.info(f"Successfully updated FASP in line log: {file_path}")
            return True

        except Exception as e:
            logging.error(f"Error updating FASP in line log: {str(e)}")
            return False
        finally:
            if wb:
                wb.close()

    @staticmethod
    def detect_range(shot_points: List[int]) -> str:
        """
        Detect ranges of consecutive shot points with 2-step intervals.

        Args:
            shot_points: List of shot point numbers

        Returns:
            Formatted string with ranges and total count

        Example:
            Input: [1001, 1003, 1005, 1011, 1013, 1015, 1017, 1031]
            Output: "Total 8 SP. 1001-1005, 1011-1017, 1031"
        """
        if not shot_points:
            return ""

        # Sort the shot points to ensure proper ordering
        sorted_points = sorted(shot_points)
        ranges = []
        start = sorted_points[0]
        end = sorted_points[0]

        for i in range(1, len(sorted_points)):
            current = sorted_points[i]
            # Check if the current point continues the sequence (2-step interval)
            if current == end + 2:
                end = current
            else:
                # End of current range, start new range
                if start == end:
                    ranges.append(str(start))
                else:
                    ranges.append(f"{start}-{end}")
                start = current
                end = current

        # Add the last range
        if start == end:
            ranges.append(str(start))
        else:
            ranges.append(f"{start}-{end}")

        total_count = len(sorted_points)
        range_str = ", ".join(ranges)

        return f"Total {total_count} SP. {range_str}"
